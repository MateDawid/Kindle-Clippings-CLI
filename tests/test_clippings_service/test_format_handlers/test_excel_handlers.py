import os
from pathlib import Path
from typing import Any
from unittest import mock

import pytest
from clippings_service.format_handlers.excel_handlers import (
    DATA_STYLING,
    FIELDS,
    HEADERS_STYLING,
    apply_data_cells_styling,
    apply_headers_styling,
    generate_excel,
)
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


@pytest.fixture
def output_excel_path(tmp_path: Path) -> str:
    """
    Returns path to output file in temporary location.

    Args:
        tmp_path (Path): Temporary pytest files location.

    Returns:
         str: Path to output file in temporary pytest files location.
    """
    return os.path.normpath(os.path.join(tmp_path, "output.xlsx"))


class TestExcelHandlers:
    """Tests for clippings_service.format_handlers.excel_handlers.py."""

    def test_apply_headers_styling(self):
        """
        GIVEN: Excel sheet with headers row.
        WHEN: Calling apply_headers_styling() function on Excel sheet.
        THEN: All headers cells filterable, styled according to HEADERS_STYLING dictionary and with width
        specified in FIELDS OrderedDict.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(list(FIELDS.keys()))

        apply_headers_styling(ws)

        assert ws.auto_filter.ref == ws.dimensions
        for cell in ws[1]:
            assert cell.font == HEADERS_STYLING["font"]
            assert cell.fill == HEADERS_STYLING["fill"]
            assert cell.alignment == HEADERS_STYLING["alignment"]
            assert cell.border == HEADERS_STYLING["border"]
        for col in ws.columns:
            col_letter = col[0].column_letter
            assert ws.column_dimensions[col_letter].width == FIELDS[col[0].value].get("width", 2)

    def test_apply_data_cells_styling(self):
        """
        GIVEN: Excel sheet with headers row and two data rows.
        WHEN: Calling apply_data_cells_styling() function on Excel sheet.
        THEN: All data cells styled according to DATA_STYLING dictionary.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(list(FIELDS.keys()))
        ws.append(
            ["Sample Book 1", "Author Name 1", "Sample content 1", "1", "1", "2025-01-01", "Highlight", "{}"],
        )
        ws.append(
            ["Sample Book 2", "Author Name 2", "Sample content 2", "2", "2", "2025-02-02", "Note", "{}"],
        )

        apply_data_cells_styling(ws)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                assert cell.font == DATA_STYLING["font"]
                assert cell.fill == DATA_STYLING["fill"]
                assert cell.alignment == DATA_STYLING["alignment"]
                assert cell.border == DATA_STYLING["border"]

    def test_generate_excel(self, output_excel_path: str, clippings_list: list[dict[str, Any]]):
        """
        GIVEN: List containing two clippings.
        WHEN: Calling generate_excel() function with clippings and output path.
        THEN: Excel file generated and containing clippings data.
        """
        result = generate_excel(clippings_list, output_excel_path)
        wb = load_workbook(output_excel_path)
        ws = wb.active

        assert result == {}
        assert os.path.exists(output_excel_path)
        assert ws.max_row == len(clippings_list) + 1
        assert ws.max_column == len(FIELDS)
        for idx, clipping in enumerate(clippings_list, start=2):
            for col in ws.columns:
                assert ws.cell(row=idx, column=col[0].col_idx).value == FIELDS[col[0].value]["fetch_method"](clipping)

    def test_generate_excel_permission_error(self, output_excel_path: str, clippings_list: list[dict[str, Any]]):
        """
        GIVEN: List containing two clippings.
        WHEN: Calling generate_excel() function with clippings and inaccessible output path.
        THEN: PermissionError raised and handled.
        """
        with mock.patch(
            "clippings_service.format_handlers.excel_handlers.Workbook.save",
            side_effect=PermissionError("Permission denied"),
        ):
            result = generate_excel(clippings_list, output_excel_path)

        assert "error" in result
        assert isinstance(result["error"], PermissionError)
        assert str(result["error"]) == "Permission denied"
