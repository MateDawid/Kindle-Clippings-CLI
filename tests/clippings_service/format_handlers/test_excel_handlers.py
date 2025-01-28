import os

import pytest
from clippings_cli.clippings_service.format_handlers.excel_handlers import (
    DATA_STYLING,
    FIELDS,
    HEADERS_STYLING,
    apply_data_cells_styling,
    apply_headers_styling,
    generate_excel,
)
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


@pytest.fixture
def output_excel_path(tmp_path, cleanup_temp_files) -> str:
    return os.path.normpath(os.path.join(tmp_path, "output.xlsx"))


class TestExcelHandlers:
    """Tests for clippings_service.format_handlers.excel_handlers.py."""

    def test_apply_headers_styling(self):
        """
        GIVEN: Excel sheet with headers row.
        WHEN: Calling apply_headers_styling() function on Excel sheet.
        THEN: All headers cells filterable, styled according to HEADERS_STYLING dictionary and with width
        specified in FIELDS OrderedDict.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(list(FIELDS.keys()))

        apply_headers_styling(ws)

        assert ws.auto_filter.ref == ws.dimensions
        for cell in ws[1]:
            assert cell.font == HEADERS_STYLING["font"]
            assert cell.fill == HEADERS_STYLING["fill"]
            assert cell.alignment == HEADERS_STYLING["alignment"]
            assert cell.border == HEADERS_STYLING["border"]
        for col in ws.columns:
            col_letter = col[0].column_letter
            assert ws.column_dimensions[col_letter].width == FIELDS[col[0].value].get("width", 2)

    def test_apply_data_cells_styling(self):
        """
        GIVEN: Excel sheet with headers row and two data rows.
        WHEN: Calling apply_data_cells_styling() function on Excel sheet.
        THEN: All data cells styled according to DATA_STYLING dictionary.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(list(FIELDS.keys()))
        ws.append(
            ["Sample Book 1", "Author Name 1", "Sample content 1", "1", "1", "2025-01-01", "Highlight", "{}"],
        )
        ws.append(
            ["Sample Book 2", "Author Name 2", "Sample content 2", "2", "2", "2025-02-02", "Note", "{}"],
        )

        apply_data_cells_styling(ws)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                assert cell.font == DATA_STYLING["font"]
                assert cell.fill == DATA_STYLING["fill"]
                assert cell.alignment == DATA_STYLING["alignment"]
                assert cell.border == DATA_STYLING["border"]

    def test_generate_excel(self, output_excel_path: str):
        """
        GIVEN: List containing two clippings.
        WHEN: Calling generate_excel() function with clippings and output path.
        THEN: Excel file generated and containing clippings data.
        """
        clippings = [
            {
                "book": {"title": "Book 1", "author": "Author 1"},
                "content": "Content 1",
                "page_number": "1",
                "location": "123",
                "created_at": "2025-01-01",
                "clipping_type": "Highlight",
                "errors": {},
            },
            {
                "book": {"title": "Book 2", "author": "Author 2"},
                "content": "Content 2",
                "page_number": "2",
                "location": "456",
                "created_at": "2025-01-02",
                "clipping_type": "Note",
                "errors": {},
            },
        ]

        result = generate_excel(clippings, output_excel_path)
        wb = load_workbook(output_excel_path)
        ws = wb.active

        assert result == {}
        assert os.path.exists(output_excel_path)
        assert ws.max_row == 3
        assert ws.max_column == len(FIELDS)
        for idx, clipping in enumerate(clippings, start=2):
            for col in ws.columns:
                assert ws.cell(row=idx, column=col[0].col_idx).value == FIELDS[col[0].value]["fetch_method"](clipping)
