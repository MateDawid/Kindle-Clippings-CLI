"""
File containing functions for handling Excel Clippings file.
"""

import os
from collections import OrderedDict

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

FIELDS: OrderedDict[str, dict] = OrderedDict(
    [
        ("Book title", {"fetch_method": lambda clipping: clipping["book"]["title"], "width": 20}),
        ("Book author", {"fetch_method": lambda clipping: clipping["book"]["author"], "width": 20}),
        ("Content", {"fetch_method": lambda clipping: clipping["content"], "width": 100}),
        ("Page number", {"fetch_method": lambda clipping: clipping["page_number"], "width": 10, "is_number": True}),
        ("Location", {"fetch_method": lambda clipping: clipping["location"], "width": 10}),
        ("Created at", {"fetch_method": lambda clipping: clipping["created_at"], "width": 10}),
        ("Clipping type", {"fetch_method": lambda clipping: clipping["clipping_type"], "width": 10}),
        ("Errors", {"fetch_method": lambda clipping: str(clipping["errors"]), "width": 20}),
    ]
)


def apply_headers_styling(ws: Worksheet) -> None:
    """
    Function to apply styling to Worksheet headers:
    * Background color
    * Font color
    * Borders
    * Enable filtering
    * Column width

    Args:
        ws (Worksheet): Excel worksheet object.
    """
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Add filters to the headers
    ws.auto_filter.ref = ws.dimensions

    # Adjust column widths
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = FIELDS[col[0].value].get("width", 2)


def apply_data_cells_styling(ws: Worksheet):
    """
    Function to apply styling to Worksheet data cells:
    * Background color
    * Font color
    * Borders

    Args:
        ws (Worksheet): Excel worksheet object.
    """
    data_font = Font(color="000000")
    data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_alignment = Alignment(horizontal="left", vertical="center")
    data_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            cell.fill = data_fill
            cell.alignment = data_alignment
            cell.border = data_border


def generate_excel(clippings: list[dict], output_path: str) -> dict:
    """
    In provided output_path creates Excel file containing data collected from Clippings input file.

    Args:
        clippings (list[dict]): List of collected Clippings.
        output_path (str): Full path to output file.

    Returns:
        dict: Dictionary containing data about potential errors.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(list(FIELDS.keys()))

    for clipping in clippings:
        ws.append([FIELDS[key]["fetch_method"](clipping) for key in FIELDS])

    apply_headers_styling(ws)
    apply_data_cells_styling(ws)

    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
    except PermissionError as e:
        return {"error": e}
    return {}
